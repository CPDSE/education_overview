"""Regenerate index.html from Pharma_DS_Course_Overview.xlsx (sheets 'SDU' and 'UCPH').

Usage: python build.py
"""
import html
import re
from datetime import date
from pathlib import Path

import openpyxl

XLSX_PATH = Path(__file__).parent / "Pharma_DS_Course_Overview.xlsx"
HTML_PATH = Path(__file__).parent / "index.html"
COLS_PER_ROW = 5

# Shared colour pool — categories are assigned slugs/colours in order, per university.
SLUG_POOL = ["cat-quant", "cat-chem", "cat-bio", "cat-physio", "cat-form", "cat-practice"]
COLOR_POOL = ["#DECBE4", "#B3CDE3", "#CCEBC5", "#FBB4AE", "#FED9A6", "#FFFFCC"]

# Fixed order preserves the SDU category->colour assignment already established.
SDU_CATEGORY_ORDER = [
    "Quantitative & Analytical Methods",
    "Chemistry",
    "Biology & Biochemistry",
    "Physiology & Pharmacology",
    "Formulation & Manufacturing",
    "Pharmacy Practice & Society",
]

UNIVERSITIES = [
    {
        "key": "sdu",
        "sheet": "SDU",
        "category_order": SDU_CATEGORY_ORDER,
        "headline_pattern": re.compile(r"Bachelor of Pharmacy – Course Overview \(SDU, [^)]*\)"),
        "headline_fmt": "Bachelor of Pharmacy – Course Overview (SDU, last updated {date})",
        "footer_pattern": re.compile(r"Last updated [^·]*· University of Southern Denmark"),
        "footer_fmt": "Last updated {date} · University of Southern Denmark",
    },
    {
        "key": "ucph",
        "sheet": "UCPH",
        "category_order": None,  # auto-detected: order of first appearance in the sheet
        "headline_pattern": re.compile(r"University of Copenhagen – Course Overview \([^)]*\)"),
        "headline_fmt": "University of Copenhagen – Course Overview (last updated {date})",
        "footer_pattern": re.compile(r"Last updated [^·]*· University of Copenhagen"),
        "footer_fmt": "Last updated {date} · University of Copenhagen",
    },
]


def fmt_ects(value):
    if value is None:
        return ""
    if float(value) == int(value):
        return f"{int(value)} ECTS"
    return f"{value}".replace(".", ",") + " ECTS"


def fmt_ds_version(value):
    if value is None or value == "":
        return ""
    if isinstance(value, str):
        return f"DS {value}"
    if float(value) == int(value):
        return f"DS v{int(value)}"
    return f"DS v{value}"


def fmt_ds_tooltip(value):
    if value is None or value == "":
        return ""
    if isinstance(value, str):
        return "Data science upgrade version not yet determined."
    version = float(value)
    if version == 1:
        return "Initial data science upgrade."
    if version == 2:
        return "Second iteration of the data science upgrade."
    if version == 3:
        return "Final version."
    if version > 2:
        return "Iterative refinement of the data science upgrade."
    return "Data science upgrade version."


def load_courses(sheet_name):
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb[sheet_name]
    headers = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}

    courses = []
    category_order = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        code = row[idx["Course Code"]]
        semester = row[idx["Semester"]]
        if code is None or semester is None:
            continue
        category = str(row[idx["Course Category 1"]] or "").strip()
        if category and category not in category_order:
            category_order.append(category)
        courses.append({
            "code": code,
            "title_da": row[idx["Danish Title"]] or "",
            "title_en": row[idx["English Title"]] or "",
            "ects": row[idx["ECTS"]],
            "semester": int(semester),
            "webpage": row[idx["Course Webpage"]] or "#",
            "category": category,
            "ds_version": row[idx["DS Version"]],
            "ds_desc": row[idx["DS Elements"]] or "",
        })

    courses.sort(key=lambda c: c["semester"])
    by_semester = {}
    for c in courses:
        by_semester.setdefault(c["semester"], []).append(c)
    return by_semester, category_order


def build_category_map(category_order):
    order = category_order[: len(SLUG_POOL)]
    slug_map = {cat: SLUG_POOL[i] for i, cat in enumerate(order)}
    color_map = {cat: COLOR_POOL[i] for i, cat in enumerate(order)}
    return order, slug_map, color_map


def render_box(course, slug_map):
    code = html.escape(str(course["code"]))
    title_da = html.escape(course["title_da"])
    title_en = html.escape(course["title_en"])
    ects = fmt_ects(course["ects"])
    webpage = html.escape(str(course["webpage"]), quote=True)
    ds_desc = html.escape(course["ds_desc"])
    link_label = "Elective catalogue" if course["code"] == "—" else "Course page"

    version_label = fmt_ds_version(course["ds_version"])
    version_tooltip = html.escape(fmt_ds_tooltip(course["ds_version"]), quote=True)
    version_tag = (
        f'<span class="tag tag-new" title="{version_tooltip}">{html.escape(version_label)}</span>'
        if version_label else ""
    )

    cat_slug = slug_map.get(course["category"], "")
    box_class = f"box {cat_slug}" if cat_slug else "box"
    bar_class = f"box-cat-bar {cat_slug}" if cat_slug else "box-cat-bar"

    return f"""    <div class="{box_class}">
      <div class="{bar_class}">{code}</div>
      <div class="box-main">
        <div class="box-title-da">{title_da}</div>
        <div class="box-title-en">{title_en}</div>
        <div class="box-meta"><span class="tag tag-ects" title="Course size in ECTS credits">{ects}</span>{version_tag}</div>
        <a class="box-link" href="{webpage}" target="_blank">&#8599; {link_label}</a>
      </div>
      <div class="box-ds"><button class="ds-toggle" aria-expanded="false" onclick="toggleDS(this)">Data Science Elements</button><div class="ds-body">{ds_desc}</div></div>
    </div>"""


def render_semester(sem_num, courses, slug_map):
    season = "Efterår" if sem_num % 2 == 1 else "Forår"
    boxes = [render_box(c, slug_map) for c in courses]
    while len(boxes) < COLS_PER_ROW:
        boxes.append('    <div class="box-empty"></div>')

    return f"""  <div class="sem-row">
    <div class="sem-label">
      <div class="sem-label-course"><div class="sem-tag">Semester</div><strong>{sem_num}</strong><span>({season})</span></div>
    </div>
{chr(10).join(boxes)}
  </div>"""


def render_grid(by_semester, slug_map):
    rows = [render_semester(sem, by_semester.get(sem, []), slug_map) for sem in sorted(by_semester)]
    return "\n\n".join(rows)


def render_legend(category_order, slug_map, color_map):
    buttons = []
    for cat in category_order:
        slug = slug_map[cat]
        color = color_map[cat]
        name = html.escape(cat)
        buttons.append(
            f'  <button class="legend-btn" data-cat="{slug}" style="background:{color}" '
            f'onclick="toggleCategoryFilter(this)">{name}</button>'
        )
    return "\n".join(buttons)


def replace_between(text, start_marker, end_marker, new_content):
    start = text.index(start_marker) + len(start_marker)
    end = text.index(end_marker, start)
    return text[:start] + "\n" + new_content + "\n" + text[end:]


def main():
    text = HTML_PATH.read_text(encoding="utf-8")
    today = date.today()
    last_updated = f"{today.day} {today.strftime('%B %Y')}"

    total_courses = 0
    for uni in UNIVERSITIES:
        by_semester, detected_order = load_courses(uni["sheet"])
        category_order = uni["category_order"] or detected_order
        category_order, slug_map, color_map = build_category_map(category_order)

        grid_html = render_grid(by_semester, slug_map)
        legend_html = render_legend(category_order, slug_map, color_map)

        text = replace_between(text, f"<!-- GRID:{uni['key']} -->", f"<!-- /GRID:{uni['key']} -->", grid_html)
        text = replace_between(text, f"<!-- LEGEND:{uni['key']} -->", f"<!-- /LEGEND:{uni['key']} -->", legend_html)

        headline = uni["headline_fmt"].format(date=last_updated)
        text = uni["headline_pattern"].sub(headline, text)

        footer_text = uni["footer_fmt"].format(date=last_updated)
        text = uni["footer_pattern"].sub(footer_text, text)

        total_courses += sum(len(v) for v in by_semester.values())

    HTML_PATH.write_text(text, encoding="utf-8")
    print(f"Wrote {HTML_PATH} from {total_courses} courses across {len(UNIVERSITIES)} universities.")


if __name__ == "__main__":
    main()
